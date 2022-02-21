# @Author  : Austin_Yee

import xlrd
import openpyxl
import re

# source=r'C:\Users\17106\Desktop\源.xls'
# template=r'C:\Users\17106\Desktop\schedule_template.xlsx'
# quantity=10
# firstSchoolDate='2020.02.21'

source_input=input(r'请输入您的源文件绝对路径（例：C:\Users\17106\Desktop\源.xls）')
template_input=input(r'请输入您的模板文件相对路径（例：C:\Users\17106\Desktop\schedule_template.xlsx）')
quantity_input=input('请输入的您课程数量（例：10）')
firstSchoolDate_input=input('请输入您的上课第一周周一的日期（例：2022.02.21）')
savePath_input=input(r'请输入您要保存的地址（例：C:\Users\17106\Desktop\）')

source=source_input
template=template_input
quantity=int(quantity_input)
firstSchoolDate=str(firstSchoolDate_input)

def getData(source,quantity):

    book = xlrd.open_workbook(source)
    s1 = book.sheet_by_index(0)
    book1 = openpyxl.load_workbook(template)
    s2 = book1.get_sheet_by_name('课程表模板')

    for i in range(0,quantity):
        # 开课班号
        classNumber = str(s1.cell(i + 1, 0).value)
        # 课程名称
        courseName = str(s1.cell(i + 1, 1).value)
        # 替换开课代码为班号
        courseName = re.sub(r'\[.*?\]', '[' + str(classNumber) + ']', courseName)
        # 开课周
        weekRange = str(s1.cell(i + 1, 5).value)+'周'
        weekRange=weekRange.replace('1 -16','1-16')
        # 班级
        classroom=str(s1.cell(i + 1,4).value)
        # 教师
        teacher=str(s1.cell(i + 1,3).value)
        # 上课节数
        whichclass=s1.row_values(i + 1,6,14) #开区间
        for j in range(0,7):
            if whichclass[j] != '':
                date1=str(s1.cell(0,j + 6).value)
                whichclass1 = whichclass[j]
                whichclass1 = whichclass1.replace('12', '1-2')

                whichclass1 = whichclass1.replace('345', '3-5')
                whichclass1 = whichclass1.replace('34', '3-4')

                whichclass1 = whichclass1.replace('6789','6-9')
                whichclass1 = whichclass1.replace('67', '6-7')

                whichclass1 = whichclass1.replace('890', '8-10')
                whichclass1 = whichclass1.replace('89', '8-9')

                whichclass1 = whichclass1.replace('ABC', '11-13')
                whichclass1 = whichclass1.replace('AB', '11-12')


                firstLine = courseName
                secondLine = weekRange + whichclass1
                thirdLine = classroom
                fourthLine = teacher

                all = firstLine + '\n' + secondLine + '\n' + thirdLine + '\n' + fourthLine
                print(all)

                # 写入空模板

                if j >0:
                    directionCell = s2.cell(row=i+2, column=j)
                    directionCell.value=all

                else:
                    directionCell = s2.cell(row=i+2, column=7)
                    directionCell.value = all
        directionCell =s2.cell(row=1,column=11)
        directionCell.value=firstSchoolDate

    book1.save(savePath_input+r'schedule_template.csv')

getData(source,quantity)
